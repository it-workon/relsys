import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

CAMINHO_PLANILHA = r"Z:\NEW TECNOLOGIA\1 Documentação\2 Inventário\6 Colaboradores Desligados\Colaboradores desligados.xlsx"
NOME_ABA = "Desligamentos"

def salvar_registro():
    try:
        wb = load_workbook(CAMINHO_PLANILHA)
        ws = wb[NOME_ABA]


        proxima_linha = ws.max_row + 1

        ws.cell(row=proxima_linha, column=1, value=colab_entry.get().upper())  # Nome colaborador
        ws.cell(row=proxima_linha, column=2, value="ok" if ad_var.get() else "")  # AD
        ws.cell(row=proxima_linha, column=3, value="ok" if email_var.get() else "")  # E-mail
        ws.cell(row=proxima_linha, column=4, value="ok" if gi_var.get() else "")  # GI
        ws.cell(row=proxima_linha, column=5, value="ok" if vbd_var.get() else "")  # VBD
        ws.cell(row=proxima_linha, column=6, value="ok" if teams_var.get() else "")  # Teams
        ws.cell(row=proxima_linha, column=7, value="ok" if forti_var.get() else "")  # FortiClient
        ws.cell(row=proxima_linha, column=8, value="ok" if papercut_var.get() else "")  # PaperCut
        ws.cell(row=proxima_linha, column=9, value=backup_entry.get())  # Backup
        ws.cell(row=proxima_linha, column=10, value=aut_entry.get())  # Autorização
        ws.cell(row=proxima_linha, column=11, value=data_entry.get())  # Data de demissão
        ws.cell(row=proxima_linha, column=12, value=zeev_entry.get())  # Zeev
        ws.cell(row=proxima_linha, column=13, value=termo_entry.get())  # Termo (link)

        wb.save(CAMINHO_PLANILHA)

        messagebox.showinfo("Sucesso", "Registro salvo na planilha com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar: {e}")

root = tk.Tk()
root.title("Checklist de Desligamento")
root.geometry("700x450")

ad_var = tk.BooleanVar()
email_var = tk.BooleanVar()
gi_var = tk.BooleanVar()
vbd_var = tk.BooleanVar()
teams_var = tk.BooleanVar()
forti_var = tk.BooleanVar()
papercut_var = tk.BooleanVar()

tk.Checkbutton(root, text="AD", variable=ad_var).grid(row=0, column=0, sticky="w")
tk.Checkbutton(root, text="E-mail", variable=email_var).grid(row=1, column=0, sticky="w")
tk.Checkbutton(root, text="GI", variable=gi_var).grid(row=2, column=0, sticky="w")
tk.Checkbutton(root, text="VBD", variable=vbd_var).grid(row=3, column=0, sticky="w")
tk.Checkbutton(root, text="Teams", variable=teams_var).grid(row=4, column=0, sticky="w")
tk.Checkbutton(root, text="FortiClient", variable=forti_var).grid(row=5, column=0, sticky="w")
tk.Checkbutton(root, text="PaperCut", variable=papercut_var).grid(row=6, column=0, sticky="w")

tk.Label(root, text="Nome do Colaborador:").grid(row=0, column=1, sticky="w")
colab_entry = tk.Entry(root, width=40)
colab_entry.grid(row=0, column=2, pady=3)

tk.Label(root, text="Backup:").grid(row=1, column=1, sticky="w")
backup_entry = tk.Entry(root, width=40)
backup_entry.grid(row=1, column=2, pady=3)

tk.Label(root, text="Autorização:").grid(row=2, column=1, sticky="w")
aut_entry = tk.Entry(root, width=40)
aut_entry.grid(row=2, column=2, pady=3)

tk.Label(root, text="Data de Demissão:").grid(row=3, column=1, sticky="w")
data_entry = tk.Entry(root, width=40)
data_entry.grid(row=3, column=2, pady=3)

tk.Label(root, text="Zeev:").grid(row=4, column=1, sticky="w")
zeev_entry = tk.Entry(root, width=40)
zeev_entry.grid(row=4, column=2, pady=3)

tk.Label(root, text="Termo (link):").grid(row=5, column=1, sticky="w")
termo_entry = tk.Entry(root, width=40)
termo_entry.grid(row=5, column=2, pady=3)

tk.Button(root, text="Salvar na Planilha", command=salvar_registro, bg="#5B7FFF", fg="white").grid(row=7, column=0, columnspan=3, pady=20)

root.mainloop()
