from openpyxl import load_workbook
from tkinter import messagebox

CAMINHO = r"Z:\NEW TECNOLOGIA\1 Documentação\2 Inventário\4 Emails\E-mails colaboradores externos - LOCAWEB.xlsx"
ABA = "BASE - LOCAWEB"

def salvar_email_externo(nome, cliente, centro_custo, descricao):

    # ---------------------------
    # 1. Validar campos
    # ---------------------------
    if not nome.strip() or not cliente.strip() or not centro_custo.strip():
        messagebox.showwarning("Campos faltando", "Nome, Cliente e Centro de Custo são obrigatórios.")
        return

    try:
        wb = load_workbook(CAMINHO)
        ws = wb[ABA]
    except Exception as e:
        messagebox.showerror("Erro", f"Não consegui abrir a planilha ou aba.\n\n{e}")
        return

    # ---------------------------
    # 2. Processar nome e gerar e-mail
    # ---------------------------
    nome_limpo = nome.strip()
    nome_completo = nome_limpo.upper()

    partes = nome_limpo.split()
    primeiro = partes[0].lower()
    ultimo = partes[-1].lower()

    email = f"{primeiro}.{ultimo}@workontrademkt.com.br"
    usuario = f"{primeiro}.{ultimo}"
    status = "ATIVO"

    # ---------------------------
    # 3. Inserir nova linha
    # ---------------------------
    linha = ws.max_row + 1

    ws.cell(row=linha, column=1, value=nome_completo)  # Nome
    ws.cell(row=linha, column=2, value=cliente)        # Cliente
    ws.cell(row=linha, column=3, value=centro_custo)   # Centro de custo
    ws.cell(row=linha, column=4, value=descricao)      # Descrição
    ws.cell(row=linha, column=5, value=email)          # E-mail
    ws.cell(row=linha, column=6, value=status)         # Status
    ws.cell(row=linha, column=7, value=usuario)        # Usuário (login)

    # ---------------------------
    # 4. Salvar planilha
    # ---------------------------
    try:
        wb.save(CAMINHO)
    except Exception as e:
        messagebox.showerror("Erro ao salvar", f"Não consegui salvar o arquivo.\n\n{e}")
        return

    # ---------------------------
    # 5. Sucesso
    # ---------------------------
    messagebox.showinfo("Sucesso", f"E-mail gerado:\n{email}")

    return CAMINHO
