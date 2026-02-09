from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side


SPREADSHEET_PATH = r"Z:\1 Documentação\2 Inventário\1 Máquinas\_Relacao_Notebooks_Desktops_atual.xlsx"
WORKSHEET_NAME = "Notebooks"


def save_machine_plan(data: list[str]) -> str:
    try:
        workbook = load_workbook(SPREADSHEET_PATH)
    except FileNotFoundError:
        workbook = Workbook()

    if WORKSHEET_NAME not in workbook.sheetnames:
        sheet = workbook.create_sheet(WORKSHEET_NAME)
        sheet.append(
            [
                "Nome do Computador",
                "Nome de Usuário",
                "Colaborador",
                "Departamento",
                "Patrimônio",
                "Locadora",
                "Modelo",
                "Office",
            ]
        )
    else:
        sheet = workbook[WORKSHEET_NAME]

    next_row = sheet.max_row + 1

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for column_index, value in enumerate(data, start=1):
        cell = sheet.cell(row=next_row, column=column_index, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    workbook.save(SPREADSHEET_PATH)
    return SPREADSHEET_PATH
