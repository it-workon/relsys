from openpyxl import load_workbook


SPREADSHEET_PATH = (
    r"Z:\NEW TECNOLOGIA\1 Documentação\2 Inventário\6 Colaboradores Desligados"
    r"\Colaboradores desligados.xlsx"
)
WORKSHEET_NAME = "Desligamentos"


def save_termination_record(values: list[str]) -> str:
    workbook = load_workbook(SPREADSHEET_PATH)
    sheet = workbook[WORKSHEET_NAME]

    next_row = sheet.max_row + 1

    for column_index, value in enumerate(values, start=1):
        sheet.cell(row=next_row, column=column_index, value=value)

    workbook.save(SPREADSHEET_PATH)
    return SPREADSHEET_PATH
