from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Border, Side

CAMINHO_PLANILHA_MAQ = r"Z:\NEW TECNOLOGIA\1 Documentação\2 Inventário\1 Máquinas\_Relacao_Notebooks_Desktops_atual.xlsx"
NOME_ABA_MAQ = "Notebooks"


def tab_plan_note(
    nome_pc,
    nome_usuario,
    colaborador,
    departamento,
    patrimonio,
    locadora,
    modelo,
    office,
):
    try:
        wb = load_workbook(CAMINHO_PLANILHA_MAQ)
    except FileNotFoundError:
        wb = Workbook()

    if NOME_ABA_MAQ not in wb.sheetnames:
        ws = wb.create_sheet(NOME_ABA_MAQ)
        ws.append(
            [
                "Nome do Computador",
                "Nome de Usuário",
                "Colaborador",
                "Departamento",
                "Patrimônio",
                "Locadora",
                "Modelo",
                "Office",
            ]
        )
    else:
        ws = wb[NOME_ABA_MAQ]

    linha = ws.max_row + 1

    valores = [
        nome_pc.upper(),
        nome_usuario,
        colaborador,
        departamento,
        patrimonio,
        locadora,
        modelo,
        office,
    ]

    borda_fina = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, valor in enumerate(valores, start=1):
        cel = ws.cell(row=linha, column=col, value=valor)
        cel.alignment = Alignment(horizontal="center", vertical="center")
        cel.border = borda_fina

    wb.save(CAMINHO_PLANILHA_MAQ)
    return CAMINHO_PLANILHA_MAQ
