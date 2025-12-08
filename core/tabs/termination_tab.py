import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import load_workbook

SHEET_PATH = r"Z:\NEW TECNOLOGIA\1 Documentação\2 Inventário\6 Colaboradores Desligados\Colaboradores desligados.xlsx"
PLANTAB_NAME = "Desligamentos"

def tab_termination(
    nome,
    ad,
    email,
    gi,
    vbd,
    teams,
    forti,
    papercut,
    backup,
    autorizacao,
    data_demissao,
    zeev,
    termo
):
    wb = load_workbook(SHEET_PATH)
    ws = wb[PLANTAB_NAME]

    linha = ws.max_row + 1

    ws.cell(row=linha, column=1, value=nome.upper())
    ws.cell(row=linha, column=2, value="ok" if ad else "")
    ws.cell(row=linha, column=3, value="ok" if email else "")
    ws.cell(row=linha, column=4, value="ok" if gi else "")
    ws.cell(row=linha, column=5, value="ok" if vbd else "")
    ws.cell(row=linha, column=6, value="ok" if teams else "")
    ws.cell(row=linha, column=7, value="ok" if forti else "")
    ws.cell(row=linha, column=8, value="ok" if papercut else "")
    ws.cell(row=linha, column=9, value=backup)
    ws.cell(row=linha, column=10, value=autorizacao)
    ws.cell(row=linha, column=11, value=data_demissao)
    ws.cell(row=linha, column=12, value=zeev)
    ws.cell(row=linha, column=13, value=termo)

    wb.save(SHEET_PATH)

    return SHEET_PATH



class TerminationTab:
    def __init__(self, app, colors):
        self.app = app
        self.colors = colors

        # Variáveis dos checkboxes
        self.ad_var = tk.BooleanVar()
        self.email_var = tk.BooleanVar()
        self.gi_var = tk.BooleanVar()
        self.vbd_var = tk.BooleanVar()
        self.teams_var = tk.BooleanVar()
        self.forti_var = tk.BooleanVar()
        self.papercut_var = tk.BooleanVar()

    def build(self, container):
        frame = ttk.Frame(container, padding=40, style="TFrame")
        frame.place(relx=0.5, rely=0.5, anchor="center")

        ttk.Label(
            frame,
            text="Checklist de Desligamento",
            style="TLabel",
            font=("Times New Roman", 16, "bold"),
            foreground=self.colors["text_color"]
        ).pack(pady=(0, 20))

        # CHECKLIST
        checks_frame = ttk.Frame(frame, style="TFrame")
        checks_frame.pack()

        col1 = ttk.Frame(checks_frame, style="TFrame")
        col1.grid(row=0, column=0, padx=25)

        col2 = ttk.Frame(checks_frame, style="TFrame")
        col2.grid(row=0, column=1, padx=25)

        ttk.Checkbutton(col1, text="AD", variable=self.ad_var, style="Checklist.TCheckbutton").pack(anchor="w")
        ttk.Checkbutton(col1, text="E-mail", variable=self.email_var, style="Checklist.TCheckbutton").pack(anchor="w")
        ttk.Checkbutton(col1, text="GI", variable=self.gi_var, style="Checklist.TCheckbutton").pack(anchor="w")
        ttk.Checkbutton(col1, text="VBD", variable=self.vbd_var, style="Checklist.TCheckbutton").pack(anchor="w")

        ttk.Checkbutton(col2, text="Teams", variable=self.teams_var, style="Checklist.TCheckbutton").pack(anchor="w")
        ttk.Checkbutton(col2, text="FortiClient", variable=self.forti_var, style="Checklist.TCheckbutton").pack(anchor="w")
        ttk.Checkbutton(col2, text="PaperCut", variable=self.papercut_var, style="Checklist.TCheckbutton").pack(anchor="w")

        # CAMPOS TEXTO
        campos_frame = ttk.Frame(frame, style="TFrame")
        campos_frame.pack(pady=20)

        def campo(nome):
            ttk.Label(campos_frame, text=nome, style="TLabel").pack(anchor="w")
            e = ttk.Entry(campos_frame, width=45)
            e.pack(pady=3)
            return e

        self.colab_entry = campo("Nome do colaborador:")
        self.backup_entry = campo("Backup:")
        self.aut_entry = campo("Autorização:")
        self.data_entry = campo("Data da demissão:")
        self.zeev_entry = campo("Zeev:")
        self.termo_entry = campo("Termo (link):")

        # BOTÃO
        ttk.Button(
            frame,
            text="Salvar na Planilha",
            style="Accent.TButton",
            command=self.on_save
        ).pack(pady=20)

        ttk.Separator(frame).pack(fill="x", pady=15)
        ttk.Label(
            frame,
            text="RelSyS © 2025",
            style="TLabel",
            font=("Times New Roman", 9, "italic"),
            foreground=self.colors["subtext_color"]
        ).pack()

    def on_save(self):
        try:
            path = tab_termination(
                self.colab_entry.get(),
                self.ad_var.get(),
                self.email_var.get(),
                self.gi_var.get(),
                self.vbd_var.get(),
                self.teams_var.get(),
                self.forti_var.get(),
                self.papercut_var.get(),
                self.backup_entry.get(),
                self.aut_entry.get(),
                self.data_entry.get(),
                self.zeev_entry.get(),
                self.termo_entry.get(),
            )

            messagebox.showinfo("Sucesso", f"Registro salvo com sucesso!\n\nArquivo:\n{path}")

        except Exception as e:
            messagebox.showerror("Erro ao salvar", str(e))
