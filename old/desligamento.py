from openpyxl import load_workbook

CAMINHO_PLANILHA = r"Z:\NEW TECNOLOGIA\1 Documentação\2 Inventário\6 Colaboradores Desligados\Colaboradores desligados.xlsx"
NOME_ABA = "Desligamentos"

def salvar_registro(
    nome,
    ad,
    email,
    gi,
    vbd,
    teams,
    forti,
    papercut,
    backup,
    autorizacao,
    data_demissao,
    zeev,
    termo
):
    wb = load_workbook(CAMINHO_PLANILHA)
    ws = wb[NOME_ABA]

    linha = ws.max_row + 1

    ws.cell(row=linha, column=1, value=nome.upper())
    ws.cell(row=linha, column=2, value="ok" if ad else "")
    ws.cell(row=linha, column=3, value="ok" if email else "")
    ws.cell(row=linha, column=4, value="ok" if gi else "")
    ws.cell(row=linha, column=5, value="ok" if vbd else "")
    ws.cell(row=linha, column=6, value="ok" if teams else "")
    ws.cell(row=linha, column=7, value="ok" if forti else "")
    ws.cell(row=linha, column=8, value="ok" if papercut else "")
    ws.cell(row=linha, column=9, value=backup)
    ws.cell(row=linha, column=10, value=autorizacao)
    ws.cell(row=linha, column=11, value=data_demissao)
    ws.cell(row=linha, column=12, value=zeev)
    ws.cell(row=linha, column=13, value=termo)

    wb.save(CAMINHO_PLANILHA)

    return CAMINHO_PLANILHA
